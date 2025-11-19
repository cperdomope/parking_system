# -*- coding: utf-8 -*-
"""
MÃ³dulo de la pestaÃ±a Reportes del sistema de gestiÃ³n de parqueadero
"""

import csv
from datetime import datetime

from PyQt5.QtCore import QDate, Qt, pyqtSignal
from PyQt5.QtWidgets import (
    QComboBox,
    QDateEdit,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from ..config.settings import CARGOS_DISPONIBLES, DIRECCIONES_DISPONIBLES
from ..database.manager import DatabaseManager

# Imports opcionales para exportaciÃ³n
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportesTab(QWidget):
    """PestaÃ±a de generaciÃ³n y visualizaciÃ³n de reportes del sistema"""

    # SeÃ±ales
    reporte_generado = pyqtSignal()

    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db = db_manager

        # Inicializar filtros sin fechas por defecto
        self.filtros_activos = {
            "tipo_vehiculo": None,
            "cargo": None,
            "direccion_grupo": None,
            "fecha_inicio": None,
            "fecha_fin": None,
        }

        self.setup_ui()
        # Actualizar reportes sin aplicar filtros inicialmente
        self.actualizar_reportes()

    def setup_ui(self):
        """Configura la interfaz de usuario de la pestaÃ±a Reportes"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 10, 15, 10)
        main_layout.setSpacing(10)

        # TÃ­tulo de la secciÃ³n
        title_label = QLabel("ðŸ“Š Sistema de Reportes")
        title_label.setStyleSheet(
            """
            font-size: 22px;
            font-weight: bold;
            color: #2c3e50;
            padding: 10px;
            background-color: #ecf0f1;
            border-radius: 5px;
        """
        )
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # SecciÃ³n de filtros
        filtros_group = QGroupBox("Filtros de Reporte")
        filtros_group.setStyleSheet(
            """
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """
        )

        # Layout vertical principal para separar en dos filas
        filtros_main_layout = QVBoxLayout()
        filtros_main_layout.setSpacing(10)

        # FILA 1: ComboBox de Tipo de VehÃ­culo, Cargo y DirecciÃ³n/Grupo
        fila1_layout = QHBoxLayout()
        fila1_layout.setSpacing(15)

        # ComboBox: Tipo de VehÃ­culo
        fila1_layout.addWidget(QLabel("Tipo VehÃ­culo:"))
        self.combo_tipo_vehiculo = QComboBox()
        self.combo_tipo_vehiculo.addItem("ðŸ“‹ Todos")
        self.combo_tipo_vehiculo.addItem("ðŸš— Carro")
        self.combo_tipo_vehiculo.addItem("ðŸï¸ Moto")
        self.combo_tipo_vehiculo.addItem("ðŸš² Bicicleta")
        self.combo_tipo_vehiculo.setStyleSheet(
            """
            QComboBox {
                padding: 5px 8px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 140px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px;
                border-left: 1px solid #bdc3c7;
            }
            QComboBox::down-arrow {
                image: none;
                width: 0;
                height: 0;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid #666;
            }
        """
        )
        fila1_layout.addWidget(self.combo_tipo_vehiculo)

        # ComboBox: Cargo
        fila1_layout.addWidget(QLabel("Cargo:"))
        self.combo_cargo = QComboBox()
        self.combo_cargo.addItem("Todos")
        self.combo_cargo.addItems(CARGOS_DISPONIBLES)
        self.combo_cargo.setStyleSheet(
            """
            QComboBox {
                padding: 5px;
                padding-right: 20px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                min-width: 140px;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #555;
                width: 0;
                height: 0;
                margin-right: 5px;
            }
        """
        )
        fila1_layout.addWidget(self.combo_cargo)

        # ComboBox: DirecciÃ³n/Grupo
        fila1_layout.addWidget(QLabel("DirecciÃ³n/Grupo:"))
        self.combo_direccion = QComboBox()
        self.combo_direccion.addItem("Todos")
        self.combo_direccion.addItems(DIRECCIONES_DISPONIBLES)
        self.combo_direccion.setStyleSheet(
            """
            QComboBox {
                padding: 5px;
                padding-right: 20px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                min-width: 200px;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #555;
                width: 0;
                height: 0;
                margin-right: 5px;
            }
        """
        )
        self.combo_direccion.view().setMinimumWidth(400)  # Para que se vean textos largos
        fila1_layout.addWidget(self.combo_direccion)
        fila1_layout.addStretch()

        # FILA 2: Fechas y Botones
        fila2_layout = QHBoxLayout()
        fila2_layout.setSpacing(15)

        # Fecha Inicial
        fila2_layout.addWidget(QLabel("Fecha Inicio:"))
        self.date_inicio = QDateEdit()
        self.date_inicio.setCalendarPopup(True)
        self.date_inicio.setDate(QDate.currentDate().addMonths(-1))  # Hace 1 mes
        self.date_inicio.setDisplayFormat("dd/MM/yyyy")
        self.date_inicio.setStyleSheet(
            """
            QDateEdit {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                min-width: 110px;
            }
        """
        )
        fila2_layout.addWidget(self.date_inicio)

        # Fecha Final
        fila2_layout.addWidget(QLabel("Fecha Fin:"))
        self.date_fin = QDateEdit()
        self.date_fin.setCalendarPopup(True)
        self.date_fin.setDate(QDate.currentDate())
        self.date_fin.setDisplayFormat("dd/MM/yyyy")
        self.date_fin.setStyleSheet(
            """
            QDateEdit {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                min-width: 110px;
            }
        """
        )
        fila2_layout.addWidget(self.date_fin)

        # BotÃ³n Limpiar Filtros
        self.btn_limpiar_filtros = QPushButton("Limpiar")
        self.btn_limpiar_filtros.setStyleSheet(
            """
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
                border-radius: 5px;
                border: none;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
            QPushButton:pressed {
                background-color: #5d6d7e;
            }
        """
        )
        self.btn_limpiar_filtros.clicked.connect(self.limpiar_filtros)
        fila2_layout.addWidget(self.btn_limpiar_filtros)

        fila2_layout.addStretch()

        # Agregar las dos filas al layout principal
        filtros_main_layout.addLayout(fila1_layout)
        filtros_main_layout.addLayout(fila2_layout)

        filtros_group.setLayout(filtros_main_layout)
        main_layout.addWidget(filtros_group)

        # Conectar filtros para actualizaciÃ³n automÃ¡tica
        self.combo_tipo_vehiculo.currentIndexChanged.connect(self.aplicar_filtros)
        self.combo_cargo.currentIndexChanged.connect(self.aplicar_filtros)
        self.combo_direccion.currentIndexChanged.connect(self.aplicar_filtros)
        self.date_inicio.dateChanged.connect(self.aplicar_filtros_fechas)
        self.date_fin.dateChanged.connect(self.aplicar_filtros_fechas)

        # TabWidget para las subpestaÃ±as de reportes
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet(
            """
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #2c3e50;
                padding: 10px 25px;
                margin-right: 2px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
                font-weight: bold;
                font-size: 12px;
            }
            QTabBar::tab:selected {
                background-color: #00B54E;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #00623A;
                color: white;
            }
        """
        )
        main_layout.addWidget(self.tab_widget)

        # Crear las 6 subpestaÃ±as
        self.tab_general = self._crear_tab_reporte_general()
        self.tab_funcionarios = self._crear_tab_funcionarios()
        self.tab_vehiculos = self._crear_tab_vehiculos()
        self.tab_parqueaderos = self._crear_tab_parqueaderos()
        self.tab_asignaciones = self._crear_tab_asignaciones()
        self.tab_excepciones = self._crear_tab_excepciones()

        self.tab_widget.addTab(self.tab_general, "ðŸ“‹ Reporte General")
        self.tab_widget.addTab(self.tab_funcionarios, "ðŸ‘¥ Funcionarios")
        self.tab_widget.addTab(self.tab_vehiculos, "ðŸš— VehÃ­culos")
        self.tab_widget.addTab(self.tab_parqueaderos, "ðŸ…¿ï¸ Parqueaderos")
        self.tab_widget.addTab(self.tab_asignaciones, "ðŸ“ Asignaciones")
        self.tab_widget.addTab(self.tab_excepciones, "ðŸ”„ Excepciones")

    def _crear_tab_reporte_general(self):
        """Crea la pestaÃ±a de Reporte General"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # DescripciÃ³n
        desc_label = QLabel("Vista consolidada del sistema: funcionarios, vehÃ­culos y ocupaciÃ³n de parqueaderos")
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic; padding: 5px;")
        layout.addWidget(desc_label)

        # Tabla
        self.tabla_general = QTableWidget()
        self.tabla_general.setColumnCount(15)
        self.tabla_general.setHorizontalHeaderLabels(
            [
                "CÃ©dula",
                "Nombre Completo",
                "Cargo",
                "DirecciÃ³n/Grupo",
                "Celular",
                "Tipo VehÃ­culo",
                "Placa",
                "CirculaciÃ³n",
                "NÂ° Parqueadero",
                "Estado Parq.",
                "Pico y Placa",
                "Discapacidad",
                "Parq. Exclusivo",
                "Carro HÃ­brido",
                "Permite Compartir",
            ]
        )
        # Configurar anchos de columnas apropiados
        self.tabla_general.setColumnWidth(0, 100)  # CÃ©dula
        self.tabla_general.setColumnWidth(1, 180)  # Nombre Completo
        self.tabla_general.setColumnWidth(2, 150)  # Cargo
        self.tabla_general.setColumnWidth(3, 150)  # DirecciÃ³n/Grupo
        self.tabla_general.setColumnWidth(4, 110)  # Celular
        self.tabla_general.setColumnWidth(5, 120)  # Tipo VehÃ­culo
        self.tabla_general.setColumnWidth(6, 90)   # Placa
        self.tabla_general.setColumnWidth(7, 100)  # CirculaciÃ³n
        self.tabla_general.setColumnWidth(8, 120)  # NÂ° Parqueadero
        self.tabla_general.setColumnWidth(9, 110)  # Estado Parq.
        self.tabla_general.setColumnWidth(10, 110) # Pico y Placa
        self.tabla_general.setColumnWidth(11, 110) # Discapacidad
        self.tabla_general.setColumnWidth(12, 120) # Parq. Exclusivo
        self.tabla_general.setColumnWidth(13, 110) # Carro HÃ­brido
        self.tabla_general.setColumnWidth(14, 130) # Permite Compartir
        self.tabla_general.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_general.horizontalHeader().setStretchLastSection(False)
        self.tabla_general.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_general.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_general.setAlternatingRowColors(True)
        self.tabla_general.setStyleSheet(
            """
            QTableWidget {
                gridline-color: #bdc3c7;
                background-color: white;
                alternate-background-color: #ecf0f1;
            }
            QHeaderView::section {
                background-color: #00B54E;
                color: white;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #009642;
            }
        """
        )
        layout.addWidget(self.tabla_general)

        # Botones de acciÃ³n
        btn_layout = self._crear_botones_exportacion(self.tabla_general, "reporte_general")
        layout.addLayout(btn_layout)

        return widget

    def _crear_tab_funcionarios(self):
        """Crea la pestaÃ±a de Funcionarios"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # DescripciÃ³n
        desc_label = QLabel("Listado completo de funcionarios activos en el sistema")
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic; padding: 5px;")
        layout.addWidget(desc_label)

        # Tabla
        self.tabla_funcionarios = QTableWidget()
        self.tabla_funcionarios.setColumnCount(10)
        self.tabla_funcionarios.setHorizontalHeaderLabels(
            [
                "ID",
                "CÃ©dula",
                "Nombre",
                "Apellidos",
                "DirecciÃ³n/Grupo",
                "Cargo",
                "Celular",
                "Tarjeta Prox.",
                "VehÃ­culos",
                "Fecha Registro",
            ]
        )
        # Configurar anchos de columnas apropiados
        self.tabla_funcionarios.setColumnWidth(0, 60)   # ID
        self.tabla_funcionarios.setColumnWidth(1, 100)  # CÃ©dula
        self.tabla_funcionarios.setColumnWidth(2, 130)  # Nombre
        self.tabla_funcionarios.setColumnWidth(3, 130)  # Apellidos
        self.tabla_funcionarios.setColumnWidth(4, 150)  # DirecciÃ³n/Grupo
        self.tabla_funcionarios.setColumnWidth(5, 150)  # Cargo
        self.tabla_funcionarios.setColumnWidth(6, 110)  # Celular
        self.tabla_funcionarios.setColumnWidth(7, 120)  # Tarjeta Prox.
        self.tabla_funcionarios.setColumnWidth(8, 90)   # VehÃ­culos
        self.tabla_funcionarios.setColumnWidth(9, 130)  # Fecha Registro
        self.tabla_funcionarios.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_funcionarios.horizontalHeader().setStretchLastSection(False)
        self.tabla_funcionarios.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_funcionarios.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_funcionarios.setAlternatingRowColors(True)
        self.tabla_funcionarios.setStyleSheet(
            """
            QTableWidget {
                gridline-color: #bdc3c7;
                background-color: white;
                alternate-background-color: #ecf0f1;
            }
            QHeaderView::section {
                background-color: #00B54E;
                color: white;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #009642;
            }
        """
        )
        layout.addWidget(self.tabla_funcionarios)

        # Botones de acciÃ³n
        btn_layout = self._crear_botones_exportacion(self.tabla_funcionarios, "reporte_funcionarios")
        layout.addLayout(btn_layout)

        return widget

    def _crear_tab_vehiculos(self):
        """Crea la pestaÃ±a de VehÃ­culos"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # DescripciÃ³n
        desc_label = QLabel("Registro de todos los vehÃ­culos y su tipo de circulaciÃ³n")
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic; padding: 5px;")
        layout.addWidget(desc_label)

        # Tabla
        self.tabla_vehiculos = QTableWidget()
        self.tabla_vehiculos.setColumnCount(9)
        self.tabla_vehiculos.setHorizontalHeaderLabels(
            [
                "ID",
                "Placa",
                "Tipo VehÃ­culo",
                "CirculaciÃ³n",
                "Propietario",
                "CÃ©dula",
                "Estado AsignaciÃ³n",
                "NÂ° Parqueadero",
                "Fecha Registro",
            ]
        )
        # Configurar anchos de columnas apropiados
        self.tabla_vehiculos.setColumnWidth(0, 60)   # ID
        self.tabla_vehiculos.setColumnWidth(1, 90)   # Placa
        self.tabla_vehiculos.setColumnWidth(2, 120)  # Tipo VehÃ­culo
        self.tabla_vehiculos.setColumnWidth(3, 110)  # CirculaciÃ³n
        self.tabla_vehiculos.setColumnWidth(4, 180)  # Propietario
        self.tabla_vehiculos.setColumnWidth(5, 100)  # CÃ©dula
        self.tabla_vehiculos.setColumnWidth(6, 150)  # Estado AsignaciÃ³n
        self.tabla_vehiculos.setColumnWidth(7, 120)  # NÂ° Parqueadero
        self.tabla_vehiculos.setColumnWidth(8, 130)  # Fecha Registro
        self.tabla_vehiculos.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_vehiculos.horizontalHeader().setStretchLastSection(False)
        self.tabla_vehiculos.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_vehiculos.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_vehiculos.setAlternatingRowColors(True)
        self.tabla_vehiculos.setStyleSheet(
            """
            QTableWidget {
                gridline-color: #bdc3c7;
                background-color: white;
                alternate-background-color: #ecf0f1;
            }
            QHeaderView::section {
                background-color: #00B54E;
                color: white;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #009642;
            }
        """
        )
        layout.addWidget(self.tabla_vehiculos)

        # Botones de acciÃ³n
        btn_layout = self._crear_botones_exportacion(self.tabla_vehiculos, "reporte_vehiculos")
        layout.addLayout(btn_layout)

        return widget

    def _crear_tab_parqueaderos(self):
        """Crea la pestaÃ±a de Parqueaderos"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # DescripciÃ³n
        desc_label = QLabel("Estado de ocupaciÃ³n de todos los parqueaderos (200 espacios en 3 sÃ³tanos)")
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic; padding: 5px;")
        layout.addWidget(desc_label)

        # Tabla
        self.tabla_parqueaderos = QTableWidget()
        self.tabla_parqueaderos.setColumnCount(8)
        self.tabla_parqueaderos.setHorizontalHeaderLabels(
            [
                "NÂ° Parqueadero",
                "SÃ³tano",
                "Tipo VehÃ­culo",
                "Estado",
                "VehÃ­culos Asignados",
                "CirculaciÃ³n PAR",
                "CirculaciÃ³n IMPAR",
                "Observaciones",
            ]
        )
        # Configurar anchos de columnas apropiados
        self.tabla_parqueaderos.setColumnWidth(0, 120)  # NÂ° Parqueadero
        self.tabla_parqueaderos.setColumnWidth(1, 80)   # SÃ³tano
        self.tabla_parqueaderos.setColumnWidth(2, 120)  # Tipo VehÃ­culo
        self.tabla_parqueaderos.setColumnWidth(3, 100)  # Estado
        self.tabla_parqueaderos.setColumnWidth(4, 150)  # VehÃ­culos Asignados
        self.tabla_parqueaderos.setColumnWidth(5, 130)  # CirculaciÃ³n PAR
        self.tabla_parqueaderos.setColumnWidth(6, 130)  # CirculaciÃ³n IMPAR
        self.tabla_parqueaderos.setColumnWidth(7, 200)  # Observaciones
        self.tabla_parqueaderos.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_parqueaderos.horizontalHeader().setStretchLastSection(False)
        self.tabla_parqueaderos.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_parqueaderos.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_parqueaderos.setAlternatingRowColors(True)
        self.tabla_parqueaderos.setStyleSheet(
            """
            QTableWidget {
                gridline-color: #bdc3c7;
                background-color: white;
                alternate-background-color: #ecf0f1;
            }
            QHeaderView::section {
                background-color: #00B54E;
                color: white;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #009642;
            }
        """
        )
        layout.addWidget(self.tabla_parqueaderos)

        # Botones de acciÃ³n
        btn_layout = self._crear_botones_exportacion(self.tabla_parqueaderos, "reporte_parqueaderos")
        layout.addLayout(btn_layout)

        return widget

    def _crear_tab_asignaciones(self):
        """Crea la pestaÃ±a de Asignaciones"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # DescripciÃ³n
        desc_label = QLabel("Asignaciones activas de vehÃ­culos a parqueaderos")
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic; padding: 5px;")
        layout.addWidget(desc_label)

        # Tabla
        self.tabla_asignaciones = QTableWidget()
        self.tabla_asignaciones.setColumnCount(10)
        self.tabla_asignaciones.setHorizontalHeaderLabels(
            [
                "ID AsignaciÃ³n",
                "NÂ° Parqueadero",
                "Placa",
                "Tipo VehÃ­culo",
                "Propietario",
                "CÃ©dula",
                "CirculaciÃ³n",
                "Fecha AsignaciÃ³n",
                "Estado",
                "Observaciones",
            ]
        )
        # Configurar anchos de columnas apropiados
        self.tabla_asignaciones.setColumnWidth(0, 100)  # ID AsignaciÃ³n
        self.tabla_asignaciones.setColumnWidth(1, 120)  # NÂ° Parqueadero
        self.tabla_asignaciones.setColumnWidth(2, 90)   # Placa
        self.tabla_asignaciones.setColumnWidth(3, 120)  # Tipo VehÃ­culo
        self.tabla_asignaciones.setColumnWidth(4, 180)  # Propietario
        self.tabla_asignaciones.setColumnWidth(5, 100)  # CÃ©dula
        self.tabla_asignaciones.setColumnWidth(6, 110)  # CirculaciÃ³n
        self.tabla_asignaciones.setColumnWidth(7, 140)  # Fecha AsignaciÃ³n
        self.tabla_asignaciones.setColumnWidth(8, 90)   # Estado
        self.tabla_asignaciones.setColumnWidth(9, 200)  # Observaciones
        self.tabla_asignaciones.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_asignaciones.horizontalHeader().setStretchLastSection(False)
        self.tabla_asignaciones.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_asignaciones.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_asignaciones.setAlternatingRowColors(True)
        self.tabla_asignaciones.setStyleSheet(
            """
            QTableWidget {
                gridline-color: #bdc3c7;
                background-color: white;
                alternate-background-color: #ecf0f1;
            }
            QHeaderView::section {
                background-color: #00B54E;
                color: white;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #009642;
            }
        """
        )
        layout.addWidget(self.tabla_asignaciones)

        # Botones de acciÃ³n
        btn_layout = self._crear_botones_exportacion(self.tabla_asignaciones, "reporte_asignaciones")
        layout.addLayout(btn_layout)

        return widget

    def _crear_tab_excepciones(self):
        """Crea la pestaÃ±a de Excepciones Pico y Placa"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(10, 10, 10, 10)

        # DescripciÃ³n
        desc_label = QLabel(
            "Funcionarios con excepciones especiales: Pico y Placa Solidario, Discapacidad, Parqueadero Exclusivo y Carro HÃ­brido"
        )
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic; padding: 5px;")
        layout.addWidget(desc_label)

        # Tabla
        self.tabla_excepciones = QTableWidget()
        self.tabla_excepciones.setColumnCount(10)
        self.tabla_excepciones.setHorizontalHeaderLabels(
            [
                "CÃ©dula",
                "Nombre Completo",
                "Cargo",
                "Pico y Placa Solidario",
                "Discapacidad",
                "Exclusivo Directivo",
                "Carro HÃ­brido",
                "Placa",
                "NÂ° Parqueadero",
                "Observaciones",
            ]
        )
        # Configurar anchos de columnas apropiados
        self.tabla_excepciones.setColumnWidth(0, 100)  # CÃ©dula
        self.tabla_excepciones.setColumnWidth(1, 180)  # Nombre Completo
        self.tabla_excepciones.setColumnWidth(2, 150)  # Cargo
        self.tabla_excepciones.setColumnWidth(3, 170)  # Pico y Placa Solidario
        self.tabla_excepciones.setColumnWidth(4, 120)  # Discapacidad
        self.tabla_excepciones.setColumnWidth(5, 150)  # Exclusivo Directivo
        self.tabla_excepciones.setColumnWidth(6, 120)  # Carro HÃ­brido
        self.tabla_excepciones.setColumnWidth(7, 90)   # Placa
        self.tabla_excepciones.setColumnWidth(8, 120)  # NÂ° Parqueadero
        self.tabla_excepciones.setColumnWidth(9, 200)  # Observaciones
        self.tabla_excepciones.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabla_excepciones.horizontalHeader().setStretchLastSection(False)
        self.tabla_excepciones.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_excepciones.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tabla_excepciones.setAlternatingRowColors(True)
        self.tabla_excepciones.setStyleSheet(
            """
            QTableWidget {
                gridline-color: #bdc3c7;
                background-color: white;
                alternate-background-color: #ecf0f1;
            }
            QHeaderView::section {
                background-color: #00B54E;
                color: white;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #009642;
            }
        """
        )
        layout.addWidget(self.tabla_excepciones)

        # Botones de acciÃ³n
        btn_layout = self._crear_botones_exportacion(self.tabla_excepciones, "reporte_excepciones")
        layout.addLayout(btn_layout)

        return widget

    def _crear_botones_exportacion(self, tabla, nombre_base):
        """Crea los botones de exportaciÃ³n para una tabla"""
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        # BotÃ³n Exportar CSV
        btn_csv = QPushButton("ðŸ“„ CSV")
        btn_csv.setStyleSheet(
            """
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 5px;
                border: none;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """
        )
        btn_csv.clicked.connect(lambda: self.exportar_csv(tabla, nombre_base))
        btn_layout.addWidget(btn_csv)

        # BotÃ³n Exportar Excel
        btn_excel = QPushButton("ðŸ“Š Excel")
        btn_excel.setStyleSheet(
            """
            QPushButton {
                background-color: #16a085;
                color: white;
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 5px;
                border: none;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #138d75;
            }
        """
        )
        btn_excel.clicked.connect(lambda: self.exportar_excel(tabla, nombre_base))
        btn_layout.addWidget(btn_excel)

        # BotÃ³n Exportar PDF
        btn_pdf = QPushButton("ðŸ“• PDF")
        btn_pdf.setStyleSheet(
            """
            QPushButton {
                background-color: #e74c3c;
                color: white;
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 5px;
                border: none;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """
        )
        btn_pdf.clicked.connect(lambda: self.exportar_pdf(tabla, nombre_base))
        btn_layout.addWidget(btn_pdf)

        btn_layout.addStretch()
        return btn_layout

    def aplicar_filtros(self):
        """Aplica los filtros seleccionados (sin fechas) y actualiza los reportes"""
        # Capturar valores de filtros (sin fechas)
        tipo_vehiculo_texto = self.combo_tipo_vehiculo.currentText()
        cargo = self.combo_cargo.currentText()
        direccion_grupo = self.combo_direccion.currentText()

        # Extraer el valor sin emoji del tipo de vehÃ­culo (ej: "ðŸš— Carro" -> "Carro")
        tipo_vehiculo = tipo_vehiculo_texto.split()[-1] if tipo_vehiculo_texto else "Todos"

        # Actualizar filtros activos (sin tocar las fechas)
        self.filtros_activos["tipo_vehiculo"] = None if tipo_vehiculo == "Todos" else tipo_vehiculo
        self.filtros_activos["cargo"] = None if cargo == "Todos" else cargo
        self.filtros_activos["direccion_grupo"] = None if direccion_grupo == "Todos" else direccion_grupo

        # Actualizar reportes con filtros
        self.actualizar_reportes()

    def aplicar_filtros_fechas(self):
        """Aplica los filtros de fechas y actualiza los reportes"""
        fecha_inicio = self.date_inicio.date().toPyDate()
        fecha_fin = self.date_fin.date().toPyDate()

        # Validar rango de fechas
        if fecha_inicio > fecha_fin:
            QMessageBox.warning(self, "Error en fechas", "La fecha de inicio no puede ser posterior a la fecha de fin.")
            return

        # Actualizar filtros de fechas
        self.filtros_activos["fecha_inicio"] = fecha_inicio
        self.filtros_activos["fecha_fin"] = fecha_fin

        # Actualizar reportes con filtros
        self.actualizar_reportes()

    def limpiar_filtros(self):
        """Limpia todos los filtros y resetea a valores por defecto"""
        # Resetear combos
        self.combo_tipo_vehiculo.setCurrentIndex(0)  # "Todos"
        self.combo_cargo.setCurrentIndex(0)  # "Todos"
        self.combo_direccion.setCurrentIndex(0)  # "Todos"

        # Resetear fechas
        self.date_inicio.setDate(QDate.currentDate().addMonths(-1))
        self.date_fin.setDate(QDate.currentDate())

        # Limpiar filtros activos
        self.filtros_activos = {
            "tipo_vehiculo": None,
            "cargo": None,
            "direccion_grupo": None,
            "fecha_inicio": None,
            "fecha_fin": None,
        }

        # Actualizar reportes sin filtros
        self.actualizar_reportes()

    def _construir_filtro_sql(self, tabla_alias=""):
        """Construye la clÃ¡usula WHERE para los filtros activos"""
        condiciones = []
        params = []

        # Prefijo para tabla
        prefix = f"{tabla_alias}." if tabla_alias else ""

        # Filtro por tipo de vehÃ­culo
        if self.filtros_activos.get("tipo_vehiculo"):
            condiciones.append(f"{prefix}tipo_vehiculo = %s")
            params.append(self.filtros_activos["tipo_vehiculo"])

        # Filtro por cargo (solo si la tabla tiene cargo)
        if self.filtros_activos.get("cargo") and "funcionario" in tabla_alias.lower():
            condiciones.append(f"{prefix}cargo = %s")
            params.append(self.filtros_activos["cargo"])

        # Filtro por rango de fechas
        if self.filtros_activos.get("fecha_inicio") and self.filtros_activos.get("fecha_fin"):
            if "fecha_registro" in tabla_alias or "fecha_creacion" in tabla_alias:
                condiciones.append(f"{prefix}fecha_registro BETWEEN %s AND %s")
                params.append(self.filtros_activos["fecha_inicio"])
                params.append(self.filtros_activos["fecha_fin"])

        return (" AND " + " AND ".join(condiciones), params) if condiciones else ("", [])

    def actualizar_reportes(self):
        """Actualiza todos los reportes cargando datos desde la base de datos"""
        errores = []

        # Actualizar cada reporte individualmente con manejo de errores
        try:
            self.actualizar_reporte_general()
        except Exception as e:
            errores.append(f"Reporte General: {str(e)}")

        try:
            self.actualizar_funcionarios()
        except Exception as e:
            errores.append(f"Funcionarios: {str(e)}")

        try:
            self.actualizar_vehiculos()
        except Exception as e:
            errores.append(f"VehÃ­culos: {str(e)}")

        try:
            self.actualizar_parqueaderos()
        except Exception as e:
            errores.append(f"Parqueaderos: {str(e)}")

        try:
            self.actualizar_asignaciones()
        except Exception as e:
            errores.append(f"Asignaciones: {str(e)}")

        try:
            self.actualizar_excepciones()
        except Exception as e:
            errores.append(f"Excepciones: {str(e)}")

        # Emitir seÃ±al de reporte generado
        self.reporte_generado.emit()

        # Mostrar resultado
        if errores:
            QMessageBox.warning(
                self, "ActualizaciÃ³n parcial", "Algunos reportes no se pudieron actualizar:\n\n" + "\n".join(errores)
            )
        else:
            # Solo mostrar mensaje de Ã©xito si se llamÃ³ desde el botÃ³n global
            # No mostrar cuando se actualiza por seÃ±ales
            pass

    def actualizar_reporte_general(self):
        """Actualiza el reporte general con datos consolidados"""
        params = []

        # Construir query con filtro de tipo de vehÃ­culo en el JOIN si es necesario
        if self.filtros_activos.get("tipo_vehiculo"):
            query = """
                SELECT
                    f.cedula,
                    CONCAT(f.nombre, ' ', f.apellidos) as nombre_completo,
                    f.cargo,
                    f.direccion_grupo,
                    f.celular,
                    v.tipo_vehiculo,
                    v.placa,
                    v.tipo_circulacion,
                    p.numero_parqueadero,
                    p.estado as estado_parqueadero,
                    CASE WHEN f.pico_placa_solidario = 1 THEN 'SÃ­' ELSE 'No' END as pico_placa_solidario,
                    CASE WHEN f.discapacidad = 1 THEN 'SÃ­' ELSE 'No' END as discapacidad,
                    CASE WHEN f.tiene_parqueadero_exclusivo = 1 THEN 'SÃ­' ELSE 'No' END as parq_exclusivo,
                    CASE WHEN f.tiene_carro_hibrido = 1 THEN 'SÃ­' ELSE 'No' END as carro_hibrido,
                    CASE WHEN f.permite_compartir = 1 THEN 'SÃ­' ELSE 'No' END as permite_compartir
                FROM funcionarios f
                LEFT JOIN vehiculos v ON f.id = v.funcionario_id AND v.activo = TRUE AND v.tipo_vehiculo = %s
                LEFT JOIN asignaciones a ON v.id = a.vehiculo_id AND a.activo = TRUE
                LEFT JOIN parqueaderos p ON a.parqueadero_id = p.id
                WHERE f.activo = TRUE
            """
            params.append(self.filtros_activos["tipo_vehiculo"])
        else:
            query = """
                SELECT
                    f.cedula,
                    CONCAT(f.nombre, ' ', f.apellidos) as nombre_completo,
                    f.cargo,
                    f.direccion_grupo,
                    f.celular,
                    v.tipo_vehiculo,
                    v.placa,
                    v.tipo_circulacion,
                    p.numero_parqueadero,
                    p.estado as estado_parqueadero,
                    CASE WHEN f.pico_placa_solidario = 1 THEN 'SÃ­' ELSE 'No' END as pico_placa_solidario,
                    CASE WHEN f.discapacidad = 1 THEN 'SÃ­' ELSE 'No' END as discapacidad,
                    CASE WHEN f.tiene_parqueadero_exclusivo = 1 THEN 'SÃ­' ELSE 'No' END as parq_exclusivo,
                    CASE WHEN f.tiene_carro_hibrido = 1 THEN 'SÃ­' ELSE 'No' END as carro_hibrido,
                    CASE WHEN f.permite_compartir = 1 THEN 'SÃ­' ELSE 'No' END as permite_compartir
                FROM funcionarios f
                LEFT JOIN vehiculos v ON f.id = v.funcionario_id AND v.activo = TRUE
                LEFT JOIN asignaciones a ON v.id = a.vehiculo_id AND a.activo = TRUE
                LEFT JOIN parqueaderos p ON a.parqueadero_id = p.id
                WHERE f.activo = TRUE
            """

        # Aplicar otros filtros
        if self.filtros_activos.get("cargo"):
            query += " AND f.cargo = %s"
            params.append(self.filtros_activos["cargo"])

        if self.filtros_activos.get("direccion_grupo"):
            query += " AND f.direccion_grupo = %s"
            params.append(self.filtros_activos["direccion_grupo"])

        if self.filtros_activos.get("fecha_inicio") and self.filtros_activos.get("fecha_fin"):
            query += " AND f.fecha_registro BETWEEN %s AND %s"
            params.append(self.filtros_activos["fecha_inicio"])
            params.append(self.filtros_activos["fecha_fin"])

        query += " ORDER BY f.apellidos, f.nombre"

        datos = self.db.fetch_all(query, tuple(params) if params else None)
        self._llenar_tabla(self.tabla_general, datos)

    def actualizar_funcionarios(self):
        """Actualiza el reporte de funcionarios"""
        # Construir query base con filtro de tipo de vehÃ­culo en el JOIN si es necesario
        params = []

        if self.filtros_activos.get("tipo_vehiculo"):
            # Si hay filtro de tipo de vehÃ­culo, aplicarlo en el JOIN
            query = """
                SELECT
                    f.id,
                    f.cedula,
                    f.nombre,
                    f.apellidos,
                    f.direccion_grupo,
                    f.cargo,
                    f.celular,
                    f.no_tarjeta_proximidad,
                    CONCAT(
                        COUNT(v.id),
                        '/',
                        CASE WHEN f.tiene_parqueadero_exclusivo = 1 THEN '6' ELSE '3' END
                    ) as total_vehiculos,
                    DATE_FORMAT(f.fecha_registro, '%Y-%m-%d') as fecha_registro
                FROM funcionarios f
                LEFT JOIN vehiculos v ON f.id = v.funcionario_id AND v.activo = TRUE AND v.tipo_vehiculo = %s
                WHERE f.activo = TRUE
            """
            params.append(self.filtros_activos["tipo_vehiculo"])
        else:
            # Sin filtro de tipo de vehÃ­culo
            query = """
                SELECT
                    f.id,
                    f.cedula,
                    f.nombre,
                    f.apellidos,
                    f.direccion_grupo,
                    f.cargo,
                    f.celular,
                    f.no_tarjeta_proximidad,
                    CONCAT(
                        COUNT(v.id),
                        '/',
                        CASE WHEN f.tiene_parqueadero_exclusivo = 1 THEN '6' ELSE '3' END
                    ) as total_vehiculos,
                    DATE_FORMAT(f.fecha_registro, '%Y-%m-%d') as fecha_registro
                FROM funcionarios f
                LEFT JOIN vehiculos v ON f.id = v.funcionario_id AND v.activo = TRUE
                WHERE f.activo = TRUE
            """

        # Aplicar otros filtros
        if self.filtros_activos.get("cargo"):
            query += " AND f.cargo = %s"
            params.append(self.filtros_activos["cargo"])

        if self.filtros_activos.get("direccion_grupo"):
            query += " AND f.direccion_grupo = %s"
            params.append(self.filtros_activos["direccion_grupo"])

        if self.filtros_activos.get("fecha_inicio") and self.filtros_activos.get("fecha_fin"):
            query += " AND f.fecha_registro BETWEEN %s AND %s"
            params.append(self.filtros_activos["fecha_inicio"])
            params.append(self.filtros_activos["fecha_fin"])

        query += """
            GROUP BY f.id, f.cedula, f.nombre, f.apellidos, f.direccion_grupo,
                     f.cargo, f.celular, f.no_tarjeta_proximidad, f.fecha_registro,
                     f.tiene_parqueadero_exclusivo
        """

        # Filtrar solo funcionarios que tengan al menos un vehÃ­culo del tipo seleccionado
        if self.filtros_activos.get("tipo_vehiculo"):
            query += " HAVING COUNT(v.id) > 0"

        query += " ORDER BY f.apellidos, f.nombre"

        datos = self.db.fetch_all(query, tuple(params) if params else None)
        self._llenar_tabla(self.tabla_funcionarios, datos)

    def actualizar_vehiculos(self):
        """Actualiza el reporte de vehÃ­culos"""
        query = """
            SELECT
                v.id,
                COALESCE(v.placa, 'Sin Placa') as placa,
                v.tipo_vehiculo,
                v.tipo_circulacion,
                CONCAT(f.nombre, ' ', f.apellidos) as propietario,
                f.cedula,
                CASE
                    WHEN a.id IS NOT NULL THEN 'Asignado'
                    ELSE 'Sin Asignar'
                END as estado_asignacion,
                COALESCE(p.numero_parqueadero, 'N/A') as numero_parqueadero,
                DATE_FORMAT(v.fecha_registro, '%Y-%m-%d') as fecha_registro
            FROM vehiculos v
            INNER JOIN funcionarios f ON v.funcionario_id = f.id
            LEFT JOIN asignaciones a ON v.id = a.vehiculo_id AND a.activo = TRUE
            LEFT JOIN parqueaderos p ON a.parqueadero_id = p.id
            WHERE v.activo = TRUE
        """

        # Aplicar filtros
        params = []
        if self.filtros_activos.get("tipo_vehiculo"):
            query += " AND v.tipo_vehiculo = %s"
            params.append(self.filtros_activos["tipo_vehiculo"])

        if self.filtros_activos.get("cargo"):
            query += " AND f.cargo = %s"
            params.append(self.filtros_activos["cargo"])

        if self.filtros_activos.get("direccion_grupo"):
            query += " AND f.direccion_grupo = %s"
            params.append(self.filtros_activos["direccion_grupo"])

        if self.filtros_activos.get("fecha_inicio") and self.filtros_activos.get("fecha_fin"):
            query += " AND v.fecha_registro BETWEEN %s AND %s"
            params.append(self.filtros_activos["fecha_inicio"])
            params.append(self.filtros_activos["fecha_fin"])

        query += " ORDER BY v.placa"

        datos = self.db.fetch_all(query, tuple(params) if params else None)
        self._llenar_tabla(self.tabla_vehiculos, datos)

    def actualizar_parqueaderos(self):
        """Actualiza el reporte de parqueaderos"""
        # Verificar si existe la columna 'sotano'
        try:
            check_query = "SHOW COLUMNS FROM parqueaderos LIKE 'sotano'"
            column_exists = self.db.fetch_one(check_query) is not None
        except Exception as e:
            print(f"Advertencia al verificar columna 'sotano': {e}")
            column_exists = False

        if column_exists:
            query = """
                SELECT
                    p.numero_parqueadero,
                    COALESCE(p.sotano, 'SÃ³tano-1') as sotano,
                    p.tipo_espacio,
                    CASE
                        -- REGLA 1: Motos y Bicicletas SIEMPRE se marcan como Completo con 1 asignaciÃ³n
                        WHEN p.tipo_espacio IN ('Moto', 'Bicicleta') AND COUNT(a.id) >= 1 THEN 'Completo'

                        -- REGLA 2: Carros con CUALQUIER tipo de excepciÃ³n se marcan como Completo desde la primera asignaciÃ³n
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 1 AND (
                            MAX(f.tiene_parqueadero_exclusivo) = 1  -- Exclusivo Directivo
                            OR MAX(f.pico_placa_solidario) = 1      -- Pico y Placa Solidario
                            OR MAX(f.discapacidad) = 1              -- Discapacidad
                            OR MAX(f.tiene_carro_hibrido) = 1       -- Carro HÃ­brido
                            OR MIN(f.permite_compartir) = 0         -- No permite compartir
                        ) THEN 'Completo'

                        -- REGLA 3: Carros regulares (sin excepciones)
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) = 1 THEN 'Parcialmente_Asignado'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 2 THEN 'Completo'

                        -- REGLA 4: Sin asignaciones
                        ELSE p.estado
                    END as estado,
                    COUNT(a.id) as vehiculos_asignados,
                    GROUP_CONCAT(CASE WHEN v.tipo_circulacion = 'PAR' THEN v.placa END) as circulacion_par,
                    GROUP_CONCAT(CASE WHEN v.tipo_circulacion = 'IMPAR' THEN v.placa END) as circulacion_impar,
                    CASE
                        -- Mismo cÃ¡lculo de estado para observaciones
                        WHEN p.tipo_espacio IN ('Moto', 'Bicicleta') AND COUNT(a.id) >= 1 THEN 'Ocupado'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 1 AND (
                            MAX(f.tiene_parqueadero_exclusivo) = 1
                            OR MAX(f.pico_placa_solidario) = 1
                            OR MAX(f.discapacidad) = 1
                            OR MAX(f.tiene_carro_hibrido) = 1
                            OR MIN(f.permite_compartir) = 0
                        ) THEN 'Ocupado'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) = 1 THEN 'Puede compartir'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 2 THEN 'Ocupado'
                        WHEN p.estado = 'Disponible' THEN 'Libre'
                        ELSE p.estado
                    END as observaciones
                FROM parqueaderos p
                LEFT JOIN asignaciones a ON p.id = a.parqueadero_id AND a.activo = TRUE
                LEFT JOIN vehiculos v ON a.vehiculo_id = v.id
                LEFT JOIN funcionarios f ON v.funcionario_id = f.id
                WHERE p.activo = TRUE
                GROUP BY p.id, p.numero_parqueadero, p.sotano, p.tipo_espacio, p.estado
                ORDER BY COALESCE(p.sotano, 'SÃ³tano-1'), p.numero_parqueadero
            """
        else:
            query = """
                SELECT
                    p.numero_parqueadero,
                    'SÃ³tano-1' as sotano,
                    p.tipo_espacio,
                    CASE
                        -- REGLA 1: Motos y Bicicletas SIEMPRE se marcan como Completo con 1 asignaciÃ³n
                        WHEN p.tipo_espacio IN ('Moto', 'Bicicleta') AND COUNT(a.id) >= 1 THEN 'Completo'

                        -- REGLA 2: Carros con CUALQUIER tipo de excepciÃ³n se marcan como Completo desde la primera asignaciÃ³n
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 1 AND (
                            MAX(f.tiene_parqueadero_exclusivo) = 1  -- Exclusivo Directivo
                            OR MAX(f.pico_placa_solidario) = 1      -- Pico y Placa Solidario
                            OR MAX(f.discapacidad) = 1              -- Discapacidad
                            OR MAX(f.tiene_carro_hibrido) = 1       -- Carro HÃ­brido
                            OR MIN(f.permite_compartir) = 0         -- No permite compartir
                        ) THEN 'Completo'

                        -- REGLA 3: Carros regulares (sin excepciones)
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) = 1 THEN 'Parcialmente_Asignado'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 2 THEN 'Completo'

                        -- REGLA 4: Sin asignaciones
                        ELSE p.estado
                    END as estado,
                    COUNT(a.id) as vehiculos_asignados,
                    GROUP_CONCAT(CASE WHEN v.tipo_circulacion = 'PAR' THEN v.placa END) as circulacion_par,
                    GROUP_CONCAT(CASE WHEN v.tipo_circulacion = 'IMPAR' THEN v.placa END) as circulacion_impar,
                    CASE
                        -- Mismo cÃ¡lculo de estado para observaciones
                        WHEN p.tipo_espacio IN ('Moto', 'Bicicleta') AND COUNT(a.id) >= 1 THEN 'Ocupado'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 1 AND (
                            MAX(f.tiene_parqueadero_exclusivo) = 1
                            OR MAX(f.pico_placa_solidario) = 1
                            OR MAX(f.discapacidad) = 1
                            OR MAX(f.tiene_carro_hibrido) = 1
                            OR MIN(f.permite_compartir) = 0
                        ) THEN 'Ocupado'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) = 1 THEN 'Puede compartir'
                        WHEN p.tipo_espacio = 'Carro' AND COUNT(a.id) >= 2 THEN 'Ocupado'
                        WHEN p.estado = 'Disponible' THEN 'Libre'
                        ELSE p.estado
                    END as observaciones
                FROM parqueaderos p
                LEFT JOIN asignaciones a ON p.id = a.parqueadero_id AND a.activo = TRUE
                LEFT JOIN vehiculos v ON a.vehiculo_id = v.id
                LEFT JOIN funcionarios f ON v.funcionario_id = f.id
                WHERE p.activo = TRUE
                GROUP BY p.id, p.numero_parqueadero, p.tipo_espacio, p.estado
                ORDER BY p.numero_parqueadero
            """

        datos = self.db.fetch_all(query)
        self._llenar_tabla(self.tabla_parqueaderos, datos)

    def actualizar_asignaciones(self):
        """Actualiza el reporte de asignaciones activas"""
        query = """
            SELECT
                a.id as id_asignacion,
                p.numero_parqueadero,
                COALESCE(v.placa, 'Sin Placa') as placa,
                v.tipo_vehiculo,
                CONCAT(f.nombre, ' ', f.apellidos) as propietario,
                f.cedula,
                v.tipo_circulacion,
                DATE_FORMAT(a.fecha_asignacion, '%Y-%m-%d %H:%i') as fecha_asignacion,
                CASE WHEN a.activo = 1 THEN 'Activa' ELSE 'Inactiva' END as estado,
                CASE
                    WHEN f.pico_placa_solidario = 1 THEN 'Pico y Placa Solidario'
                    WHEN f.tiene_parqueadero_exclusivo = 1 THEN 'Exclusivo Directivo'
                    ELSE 'Normal'
                END as observaciones
            FROM asignaciones a
            INNER JOIN parqueaderos p ON a.parqueadero_id = p.id
            INNER JOIN vehiculos v ON a.vehiculo_id = v.id
            INNER JOIN funcionarios f ON v.funcionario_id = f.id
            WHERE a.activo = TRUE
        """

        # Aplicar filtros
        params = []
        if self.filtros_activos.get("tipo_vehiculo"):
            query += " AND v.tipo_vehiculo = %s"
            params.append(self.filtros_activos["tipo_vehiculo"])

        if self.filtros_activos.get("cargo"):
            query += " AND f.cargo = %s"
            params.append(self.filtros_activos["cargo"])

        if self.filtros_activos.get("direccion_grupo"):
            query += " AND f.direccion_grupo = %s"
            params.append(self.filtros_activos["direccion_grupo"])

        if self.filtros_activos.get("fecha_inicio") and self.filtros_activos.get("fecha_fin"):
            query += " AND a.fecha_asignacion BETWEEN %s AND %s"
            params.append(self.filtros_activos["fecha_inicio"])
            params.append(self.filtros_activos["fecha_fin"])

        query += " ORDER BY a.fecha_asignacion DESC"

        datos = self.db.fetch_all(query, tuple(params) if params else None)
        self._llenar_tabla(self.tabla_asignaciones, datos)

    def actualizar_excepciones(self):
        """Actualiza el reporte de excepciones (Pico y Placa Solidario, Discapacidad, Exclusivo, Carro HÃ­brido)"""
        params = []

        # Construir query con filtro de tipo de vehÃ­culo en el JOIN si es necesario
        if self.filtros_activos.get("tipo_vehiculo"):
            query = """
                SELECT
                    f.cedula,
                    CONCAT(f.nombre, ' ', f.apellidos) as nombre_completo,
                    f.cargo,
                    CASE WHEN f.pico_placa_solidario = 1 THEN 'SI' ELSE 'NO' END as pico_placa_solidario,
                    CASE WHEN f.discapacidad = 1 THEN 'SI' ELSE 'NO' END as discapacidad,
                    CASE WHEN f.tiene_parqueadero_exclusivo = 1 THEN 'SI' ELSE 'NO' END as parqueadero_exclusivo,
                    CASE WHEN f.tiene_carro_hibrido = 1 THEN 'SI' ELSE 'NO' END as carro_hibrido,
                    COALESCE(v.placa, 'Sin vehÃ­culo') as placa,
                    COALESCE(p.numero_parqueadero, 'N/A') as numero_parqueadero,
                    CASE
                        WHEN f.pico_placa_solidario = 1 THEN 'Puede usar parqueadero cualquier dÃ­a'
                        WHEN f.discapacidad = 1 THEN 'Prioridad para espacios especiales'
                        WHEN f.tiene_parqueadero_exclusivo = 1 THEN 'Parqueadero exclusivo (hasta 4 vehÃ­culos)'
                        WHEN f.tiene_carro_hibrido = 1 THEN 'Carro hÃ­brido - Sin restricciÃ³n pico y placa'
                        ELSE 'Sin excepciones'
                    END as observaciones
                FROM funcionarios f
                LEFT JOIN vehiculos v ON f.id = v.funcionario_id AND v.activo = TRUE AND v.tipo_vehiculo = %s
                LEFT JOIN asignaciones a ON v.id = a.vehiculo_id AND a.activo = TRUE
                LEFT JOIN parqueaderos p ON a.parqueadero_id = p.id
                WHERE f.activo = TRUE
                    AND (f.pico_placa_solidario = 1 OR f.discapacidad = 1 OR f.tiene_parqueadero_exclusivo = 1 OR f.tiene_carro_hibrido = 1)
            """
            params.append(self.filtros_activos["tipo_vehiculo"])
        else:
            query = """
                SELECT
                    f.cedula,
                    CONCAT(f.nombre, ' ', f.apellidos) as nombre_completo,
                    f.cargo,
                    CASE WHEN f.pico_placa_solidario = 1 THEN 'SI' ELSE 'NO' END as pico_placa_solidario,
                    CASE WHEN f.discapacidad = 1 THEN 'SI' ELSE 'NO' END as discapacidad,
                    CASE WHEN f.tiene_parqueadero_exclusivo = 1 THEN 'SI' ELSE 'NO' END as parqueadero_exclusivo,
                    CASE WHEN f.tiene_carro_hibrido = 1 THEN 'SI' ELSE 'NO' END as carro_hibrido,
                    COALESCE(v.placa, 'Sin vehÃ­culo') as placa,
                    COALESCE(p.numero_parqueadero, 'N/A') as numero_parqueadero,
                    CASE
                        WHEN f.pico_placa_solidario = 1 THEN 'Puede usar parqueadero cualquier dÃ­a'
                        WHEN f.discapacidad = 1 THEN 'Prioridad para espacios especiales'
                        WHEN f.tiene_parqueadero_exclusivo = 1 THEN 'Parqueadero exclusivo (hasta 4 vehÃ­culos)'
                        WHEN f.tiene_carro_hibrido = 1 THEN 'Carro hÃ­brido - Sin restricciÃ³n pico y placa'
                        ELSE 'Sin excepciones'
                    END as observaciones
                FROM funcionarios f
                LEFT JOIN vehiculos v ON f.id = v.funcionario_id AND v.activo = TRUE
                LEFT JOIN asignaciones a ON v.id = a.vehiculo_id AND a.activo = TRUE
                LEFT JOIN parqueaderos p ON a.parqueadero_id = p.id
                WHERE f.activo = TRUE
                    AND (f.pico_placa_solidario = 1 OR f.discapacidad = 1 OR f.tiene_parqueadero_exclusivo = 1 OR f.tiene_carro_hibrido = 1)
            """

        # Aplicar otros filtros
        if self.filtros_activos.get("cargo"):
            query += " AND f.cargo = %s"
            params.append(self.filtros_activos["cargo"])

        if self.filtros_activos.get("direccion_grupo"):
            query += " AND f.direccion_grupo = %s"
            params.append(self.filtros_activos["direccion_grupo"])

        if self.filtros_activos.get("fecha_inicio") and self.filtros_activos.get("fecha_fin"):
            query += " AND f.fecha_registro BETWEEN %s AND %s"
            params.append(self.filtros_activos["fecha_inicio"])
            params.append(self.filtros_activos["fecha_fin"])

        query += " ORDER BY f.apellidos, f.nombre"

        datos = self.db.fetch_all(query, tuple(params) if params else None)
        self._llenar_tabla(self.tabla_excepciones, datos)

    def _llenar_tabla(self, tabla, datos):
        """Llena una tabla con los datos proporcionados"""
        tabla.setRowCount(0)

        if not datos:
            return

        for row_data in datos:
            row_position = tabla.rowCount()
            tabla.insertRow(row_position)

            for col_index, (key, value) in enumerate(row_data.items()):
                item = QTableWidgetItem(str(value) if value is not None else "")
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Solo lectura
                tabla.setItem(row_position, col_index, item)

    def exportar_csv(self, tabla, nombre_base):
        """Exporta los datos de la tabla a un archivo CSV"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename, _ = QFileDialog.getSaveFileName(
                self, "Guardar como CSV", f"{nombre_base}_{timestamp}.csv", "Archivos CSV (*.csv)"
            )

            if not filename:
                return

            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)

                # Escribir encabezados
                headers = []
                for col in range(tabla.columnCount()):
                    headers.append(tabla.horizontalHeaderItem(col).text())
                writer.writerow(headers)

                # Escribir datos
                for row in range(tabla.rowCount()):
                    row_data = []
                    for col in range(tabla.columnCount()):
                        item = tabla.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)

            QMessageBox.information(self, "Ã‰xito", f"Datos exportados correctamente a:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar CSV: {str(e)}")

    def exportar_excel(self, tabla, nombre_base):
        """Exporta los datos a Excel usando openpyxl"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                self,
                "LibrerÃ­a no disponible",
                "La exportaciÃ³n a Excel requiere instalar la librerÃ­a 'openpyxl'.\n\n"
                "Instalar con: pip install openpyxl\n\n"
                "Por ahora, use la exportaciÃ³n a CSV.",
            )
            return

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename, _ = QFileDialog.getSaveFileName(
                self, "Guardar como Excel", f"{nombre_base}_{timestamp}.xlsx", "Archivos Excel (*.xlsx)"
            )

            if not filename:
                return

            # Crear workbook y hoja activa
            wb = Workbook()
            ws = wb.active
            ws.title = nombre_base[:31]  # Excel limita nombres de hoja a 31 caracteres

            # Estilo para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Escribir encabezados
            headers = []
            for col in range(tabla.columnCount()):
                header_text = tabla.horizontalHeaderItem(col).text()
                headers.append(header_text)
                cell = ws.cell(row=1, column=col + 1, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Escribir datos
            for row in range(tabla.rowCount()):
                for col in range(tabla.columnCount()):
                    item = tabla.item(row, col)
                    value = item.text() if item else ""
                    cell = ws.cell(row=row + 2, column=col + 1, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Ajustar ancho de columnas
            for col in range(1, tabla.columnCount() + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

            # Guardar archivo
            wb.save(filename)
            QMessageBox.information(self, "Ã‰xito", f"Datos exportados correctamente a:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar Excel: {str(e)}")

    def exportar_pdf(self, tabla, nombre_base):
        """Exporta los datos a PDF usando reportlab"""
        if not REPORTLAB_AVAILABLE:
            QMessageBox.warning(
                self,
                "LibrerÃ­a no disponible",
                "La exportaciÃ³n a PDF requiere instalar la librerÃ­a 'reportlab'.\n\n"
                "Instalar con: pip install reportlab\n\n"
                "Por ahora, use la exportaciÃ³n a CSV.",
            )
            return

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename, _ = QFileDialog.getSaveFileName(
                self, "Guardar como PDF", f"{nombre_base}_{timestamp}.pdf", "Archivos PDF (*.pdf)"
            )

            if not filename:
                return

            # Crear documento PDF
            doc = SimpleDocTemplate(
                filename, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
            )

            # Contenedor de elementos
            elements = []
            styles = getSampleStyleSheet()

            # TÃ­tulo del reporte
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                textColor=colors.HexColor("#2c3e50"),
                spaceAfter=20,
                alignment=1,  # Centrado
            )
            title = Paragraph(f"Reporte: {nombre_base.replace('_', ' ').title()}", title_style)
            elements.append(title)

            # Fecha y hora de generaciÃ³n
            date_style = ParagraphStyle(
                "DateStyle",
                parent=styles["Normal"],
                fontSize=10,
                textColor=colors.HexColor("#7f8c8d"),
                spaceAfter=20,
                alignment=1,
            )
            date_text = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", date_style)
            elements.append(date_text)
            elements.append(Spacer(1, 0.2 * inch))

            # Extraer datos de la tabla
            data = []

            # Encabezados
            headers = []
            for col in range(tabla.columnCount()):
                headers.append(tabla.horizontalHeaderItem(col).text())
            data.append(headers)

            # Filas de datos
            for row in range(tabla.rowCount()):
                row_data = []
                for col in range(tabla.columnCount()):
                    item = tabla.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)

            # Crear tabla PDF
            # Calcular ancho de columnas dinÃ¡micamente
            page_width = landscape(A4)[0] - 60  # Restar mÃ¡rgenes
            col_width = page_width / tabla.columnCount()

            table = Table(data, colWidths=[col_width] * tabla.columnCount())

            # Estilo de la tabla
            table.setStyle(
                TableStyle(
                    [
                        # Encabezados
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        # Datos
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )

            elements.append(table)

            # Pie de pÃ¡gina con informaciÃ³n del sistema
            elements.append(Spacer(1, 0.3 * inch))
            footer_style = ParagraphStyle(
                "FooterStyle", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#95a5a6"), alignment=1
            )
            footer = Paragraph("Sistema de GestiÃ³n de Parqueadero - Ssalud Plaza Claro Â© 2025", footer_style)
            elements.append(footer)

            # Construir PDF
            doc.build(elements)
            QMessageBox.information(self, "Ã‰xito", f"PDF generado correctamente en:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar PDF: {str(e)}")
